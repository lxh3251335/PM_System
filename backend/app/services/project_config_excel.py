"""
项目配置 Excel 导出 / 导入（多工作表，与 demo 全量配置对齐）
"""
from __future__ import annotations

import enum as py_enum
import io
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple, Type, TypeVar

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..models.device import Device, DeviceRelation, DeviceType, RelationType
from ..models.project import ColdRoom, ColdRoomType, Project, ProjectStatus


EXPORT_VERSION = 2

PROJECT_KV_FIELDS = [
    ("project_id", "项目ID（导入时须与目标项目一致）"),
    ("project_no", "项目编号（只读，勿改）"),
    ("name", "项目名称"),
    ("end_customer", "最终用户"),
    ("business_type", "业务类型"),
    ("city", "城市"),
    ("address", "项目地址"),
    ("mailing_address", "邮寄地址"),
    ("recipient_name", "收件人"),
    ("recipient_phone", "收件电话"),
    ("expected_arrival_time", "期望到货时间(YYYY-MM-DD)"),
    ("status", "状态"),
    ("remarks", "备注"),
]

PROJECT_KV_ENGLISH_KEYS = frozenset(k for k, _ in PROJECT_KV_FIELDS)

# 横向表头（第 1 行）常见中文/英文 -> 内部字段名
HEADER_TO_PROJECT_KEY: Dict[str, str] = {}
for _k, _ in PROJECT_KV_FIELDS:
    HEADER_TO_PROJECT_KEY[_k] = _k
    HEADER_TO_PROJECT_KEY[_k.casefold()] = _k
_EXTRA_HEADER_MAP = {
    "项目id": "project_id",
    "项目编号": "project_no",
    "项目编号（只读）": "project_no",
    "项目编号(只读)": "project_no",
    "项目名称": "name",
    "项目名": "name",
    "城市": "city",
    "业务类型": "business_type",
    "地址": "address",
    "收件人": "recipient_name",
    "备注": "remarks",
    "状态": "status",
    "终端客户": "end_customer",
    "最终用户": "end_customer",
    "项目地址": "address",
    "详细地址": "address",
    "邮寄地址": "mailing_address",
    "收件人姓名": "recipient_name",
    "联系电话": "recipient_phone",
    "收件电话": "recipient_phone",
    "电话": "recipient_phone",
    "到货时间": "expected_arrival_time",
    "期望到货": "expected_arrival_time",
    "流程状态": "status",
    "工程名称": "name",
    "工程名": "name",
    "工程编号": "project_no",
    "项目代号": "project_no",
    "项目代码": "project_no",
    "客户名称": "end_customer",
    "客户": "end_customer",
    "项目 ID": "project_id",
}
for _cn, _en in _EXTRA_HEADER_MAP.items():
    HEADER_TO_PROJECT_KEY[_cn] = _en
    HEADER_TO_PROJECT_KEY[_cn.casefold()] = _en

# 竖向「项目」表：中文列 A 与内部字段名互认；横向表头亦识别完整中文说明列
for _ek, _lab in PROJECT_KV_FIELDS:
    HEADER_TO_PROJECT_KEY[_lab] = _ek

COLD_HEADERS = [
    "冷库名称",
    "冷库类型",
    "设计温度下限",
    "设计温度上限",
    "面积(㎡)",
    "高度(m)",
    "容积(m³)",
    "制冷剂类型",
]

DEVICE_HEADERS = [
    "设备编号",
    "设备类型",
    "所属冷库名称",
    "品牌",
    "型号",
    "融霜方式",
    "智能融霜",
    "膨胀阀方式",
    "出厂编号",
    "通讯端口类型",
    "通讯协议",
    "网关编号",
    "网关端口",
    "RS485通讯地址",
    "规格参数（JSON）",
    "备注",
]

REL_HEADERS = ["源设备编号", "目标设备编号", "关系类型", "描述"]

# 冷库类型：导出中文 / 导入兼容英文枚举值与中文
COLD_ROOM_TYPE_ZH: Dict[ColdRoomType, str] = {
    ColdRoomType.LOW_TEMP: "低温库",
    ColdRoomType.REFRIGERATED: "冷藏库",
    ColdRoomType.MEDIUM_TEMP: "中温库",
}
COLD_ROOM_STR_TO_TYPE: Dict[str, ColdRoomType] = {}
for _e, _zh in COLD_ROOM_TYPE_ZH.items():
    COLD_ROOM_STR_TO_TYPE[_zh] = _e
    COLD_ROOM_STR_TO_TYPE[_e.value] = _e
    COLD_ROOM_STR_TO_TYPE[_e.value.casefold()] = _e
for _e in ColdRoomType:
    COLD_ROOM_STR_TO_TYPE[_e.name] = _e
    COLD_ROOM_STR_TO_TYPE[_e.name.upper()] = _e

# 设备类型
DEVICE_TYPE_ZH: Dict[DeviceType, str] = {
    DeviceType.AIR_COOLER: "冷风机",
    DeviceType.THERMOSTAT: "温控器",
    DeviceType.UNIT: "机组",
    DeviceType.METER: "电表",
    DeviceType.FREEZER: "冷柜",
    DeviceType.DEFROST_CONTROLLER: "智能融霜控制器",
}
DEVICE_STR_TO_TYPE: Dict[str, DeviceType] = {}
for _e, _zh in DEVICE_TYPE_ZH.items():
    DEVICE_STR_TO_TYPE[_zh] = _e
    DEVICE_STR_TO_TYPE[_e.value] = _e
    DEVICE_STR_TO_TYPE[_e.value.casefold()] = _e
for _e in DeviceType:
    DEVICE_STR_TO_TYPE[_e.name] = _e
    DEVICE_STR_TO_TYPE[_e.name.upper()] = _e

# 库表未建模但前端/历史数据可能出现的类型（仅导出中文展示）
DEVICE_TYPE_ORPHAN_ZH: Dict[str, str] = {
    "cabinet": "电控柜",
    "CABINET": "电控柜",
    "diankong": "电控柜",
    "DIANKONG": "电控柜",
    "cooling_tower": "冷却塔",
    "COOLING_TOWER": "冷却塔",
}

# 设备关系类型
RELATION_TYPE_ZH: Dict[RelationType, str] = {
    RelationType.THERMOSTAT_TO_AIR_COOLER: "温控器→冷风机",
    RelationType.UNIT_TO_AIR_COOLER: "机组→冷风机",
    RelationType.UNIT_TO_FREEZER: "机组→冷柜",
    RelationType.METER_TO_UNIT: "电表→机组",
    RelationType.METER_TO_AIR_COOLER: "电表→冷风机",
    RelationType.DEFROST_TO_AIR_COOLER: "融霜控制器→冷风机",
}
RELATION_STR_TO_TYPE: Dict[str, RelationType] = {}
for _e, _zh in RELATION_TYPE_ZH.items():
    RELATION_STR_TO_TYPE[_zh] = _e
    RELATION_STR_TO_TYPE[_zh.replace("→", "->")] = _e
    RELATION_STR_TO_TYPE[_e.value] = _e
    RELATION_STR_TO_TYPE[_e.value.casefold()] = _e
for _e in RelationType:
    RELATION_STR_TO_TYPE[_e.name] = _e
    RELATION_STR_TO_TYPE[_e.name.upper()] = _e

PROJECT_STATUS_ZH: Dict[ProjectStatus, str] = {
    ProjectStatus.NEW: "新项目",
    ProjectStatus.EQUIPMENT_REGISTRATION: "设备登记中",
    ProjectStatus.SUBMITTED: "已提交待配置",
    ProjectStatus.CONFIGURING: "管理员配置中",
    ProjectStatus.COMPLETED: "已完成",
}

LABEL_TO_PROJECT_KEY: Dict[str, str] = {}
for _ek, _lab in PROJECT_KV_FIELDS:
    LABEL_TO_PROJECT_KEY[_lab] = _ek
    LABEL_TO_PROJECT_KEY[_lab.strip().replace("\u3000", "").replace("\xa0", "")] = _ek

E = TypeVar("E", bound=py_enum.Enum)


def _resolve_stored_enum(enum_cls: Type[E], raw: Any) -> Optional[E]:
    """
    ORM 枚举实例 / 枚举 value 字符串 / SQLite 中常见的枚举成员名（如 AIR_COOLER）统一解析。
    """
    if raw is None:
        return None
    if isinstance(raw, enum_cls):
        return raw
    s = str(getattr(raw, "value", raw)).strip()
    if not s:
        return None
    try:
        return enum_cls(s)  # type: ignore[return-value]
    except ValueError:
        pass
    sup = s.upper().replace("-", "_")
    try:
        return enum_cls[sup]  # type: ignore[index]
    except KeyError:
        pass
    for m in enum_cls:
        if m.name == sup or m.value == s or str(m.value).lower() == s.lower():
            return m  # type: ignore[return-value]
    return None


def _room_type_label(rt: Any) -> str:
    """导出用中文冷库类型（导入仍兼容英文枚举值）"""
    e = _resolve_stored_enum(ColdRoomType, rt)
    if e is not None:
        return COLD_ROOM_TYPE_ZH.get(e, e.value)
    s = str(getattr(rt, "value", rt))
    return s


def _dt_label(dt: Any) -> str:
    """导出用中文设备类型"""
    e = _resolve_stored_enum(DeviceType, dt)
    if e is not None:
        return DEVICE_TYPE_ZH.get(e, e.value)
    s = str(getattr(dt, "value", dt)).strip()
    return DEVICE_TYPE_ORPHAN_ZH.get(s, DEVICE_TYPE_ORPHAN_ZH.get(s.upper(), s))


def _relation_type_label(rt: Any) -> str:
    """导出用中文关系类型"""
    e = _resolve_stored_enum(RelationType, rt)
    if e is not None:
        return RELATION_TYPE_ZH.get(e, e.value)
    s = str(getattr(rt, "value", rt))
    return s


def _project_status_zh(st: Any) -> str:
    e = _resolve_stored_enum(ProjectStatus, st)
    if e is not None:
        return PROJECT_STATUS_ZH.get(e, str(e.value))
    s = str(getattr(st, "value", st))
    return s


def _cold_room_name(device: Device, room_by_id: Dict[int, str]) -> str:
    if not device.cold_room_id:
        return ""
    return room_by_id.get(device.cold_room_id, "")


def build_workbook_bytes(db: Session, project_id: int) -> Tuple[bytes, str]:
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise ValueError("项目不存在")

    cold_rooms: List[ColdRoom] = (
        db.query(ColdRoom).filter(ColdRoom.project_id == project_id).order_by(ColdRoom.id).all()
    )
    room_by_id = {r.id: r.name for r in cold_rooms}
    devices: List[Device] = db.query(Device).filter(Device.project_id == project_id).all()

    def _dev_export_sort_key(d: Device) -> Tuple[str, str, int]:
        rn = room_by_id.get(d.cold_room_id, "") if d.cold_room_id else ""
        dv = d.device_type.value if hasattr(d.device_type, "value") else str(d.device_type)
        return (rn, dv, d.id)

    devices = sorted(devices, key=_dev_export_sort_key)
    relations: List[DeviceRelation] = (
        db.query(DeviceRelation)
        .filter(DeviceRelation.project_id == project_id)
        .order_by(DeviceRelation.id)
        .all()
    )
    dev_by_id = {d.id: d for d in devices}

    wb = Workbook()
    # 说明
    ws0 = wb.active
    ws0.title = "说明"
    ws0["A1"] = "冷库项目配置导出"
    ws0["A2"] = f"导出版本: {EXPORT_VERSION}"
    ws0["A3"] = "导入：在「项目管理」列表对该行项目点「导入配置」，选择本文件。"
    ws0["A4"] = "请勿改工作表名称；「项目」表第一列为中文字段名，其中「项目ID」须与导入目标项目一致。"
    ws0["A5"] = "提示：「导出配置」下载的是本系统生成的中文模板；「下载附件」为您此前上传的文件，语言/版式可能与当前导出不一致。"
    ws0["A6"] = f"导出版本号 {EXPORT_VERSION}：业务表（项目/冷库/设备/设备关系）为中文表头与中文类型名；「导出信息」表仅作版本标记。"

    ws_meta = wb.create_sheet("导出信息")
    ws_meta["A1"] = "导出版本"
    ws_meta["B1"] = EXPORT_VERSION

    ws_p = wb.create_sheet("项目")
    ws_p.append(["字段", "值"])
    values: Dict[str, Any] = {
        "project_id": project.id,
        "project_no": project.project_no or "",
        "name": project.name or "",
        "end_customer": project.end_customer or "",
        "business_type": project.business_type or "",
        "city": project.city or "",
        "address": project.address or "",
        "mailing_address": project.mailing_address or "",
        "recipient_name": project.recipient_name or "",
        "recipient_phone": project.recipient_phone or "",
        "expected_arrival_time": project.expected_arrival_time.isoformat()
        if project.expected_arrival_time
        else "",
        "status": _project_status_zh(project.status),
        "remarks": project.remarks or "",
    }
    for key, label in PROJECT_KV_FIELDS:
        ws_p.append([label, values.get(key, "")])

    ws_c = wb.create_sheet("冷库")
    ws_c.append(COLD_HEADERS)
    for r in cold_rooms:
        ws_c.append(
            [
                r.name,
                _room_type_label(r.room_type),
                r.design_temp_min,
                r.design_temp_max,
                r.area,
                r.height,
                r.volume,
                r.refrigerant_type or "",
            ]
        )

    ws_d = wb.create_sheet("设备")
    ws_d.append(DEVICE_HEADERS)
    for d in devices:
        ws_d.append(
            [
                d.device_no,
                _dt_label(d.device_type),
                _cold_room_name(d, room_by_id),
                d.brand,
                d.model,
                d.defrost_method or "",
                d.has_intelligent_defrost or "",
                d.expansion_valve_type or "",
                d.factory_no or "",
                d.comm_port_type or "",
                d.comm_protocol or "",
                d.gateway_id or "",
                d.gateway_port or "",
                d.rs485_address or "",
                d.specifications or "",
                d.remarks or "",
            ]
        )

    ws_r = wb.create_sheet("设备关系")
    ws_r.append(REL_HEADERS)
    for rel in relations:
        fd = dev_by_id.get(rel.from_device_id)
        td = dev_by_id.get(rel.to_device_id)
        ws_r.append(
            [
                fd.device_no if fd else rel.from_device_id,
                td.device_no if td else rel.to_device_id,
                _relation_type_label(rel.relation_type),
                rel.description or "",
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = (project.project_no or f"project_{project_id}").replace("/", "-")
    return buf.getvalue(), safe_name


def _resolve_header_to_project_key(hs: str) -> Optional[str]:
    """表头文案 -> 内部项目字段（支持空格、全角括号差异）"""
    if not hs:
        return None
    hs = _norm_sheet_title(hs)
    if not hs:
        return None
    variants = [
        hs,
        hs.casefold(),
        re.sub(r"\s+", "", hs),
        re.sub(r"\s+", "", hs).casefold(),
    ]
    for v in variants:
        k = HEADER_TO_PROJECT_KEY.get(v) or HEADER_TO_PROJECT_KEY.get(v.casefold())
        if k:
            return k
    hsf = hs.casefold().replace(" ", "_").replace("（", "(").replace("）", ")")
    if hsf in PROJECT_KV_ENGLISH_KEYS:
        return hsf
    return None


def _sheet_has_resolvable_project_header(ws) -> bool:
    """前若干行内是否出现可映射到项目字段的表头文案（兼容默认表名 Sheet / 合并单元格旁移）"""
    try:
        for row in ws.iter_rows(min_row=1, max_row=45, max_col=60, values_only=True):
            if not row:
                continue
            for cell in row:
                if cell is None:
                    continue
                if isinstance(cell, (int, float)) and not isinstance(cell, bool):
                    continue
                hs = _norm_sheet_title(str(cell))
                if hs and _resolve_header_to_project_key(hs):
                    return True
    except Exception:
        return False
    return False


def _try_wide_project_kv_from_rows(rows: List[Any]) -> Optional[Dict[str, Any]]:
    """
    横向布局：第 1 行为表头（字段名），第 2 行为对应取值。
    与竖向「字段|值」两列表区分：竖向首格常为「字段」或首行仅 2 列且第二列表头是数值。
    """
    if len(rows) < 2:
        return None
    header_row = rows[0]
    value_row = rows[1]
    if not header_row or not value_row:
        return None
    if header_row[0] is not None:
        a0 = _norm_sheet_title(str(header_row[0])).casefold()
        if a0 in ("字段", "field", "key", "项目字段"):
            return None

    string_headers = 0
    mapped = 0
    out: Dict[str, Any] = {}
    max_i = max(len(header_row), len(value_row))
    for i in range(max_i):
        hcell = header_row[i] if i < len(header_row) else None
        if hcell is None:
            continue
        if isinstance(hcell, (int, float)) and not isinstance(hcell, bool):
            continue
        hs = _norm_sheet_title(str(hcell))
        if not hs:
            continue
        string_headers += 1
        key = _resolve_header_to_project_key(hs)
        if key is None:
            continue
        val = value_row[i] if i < len(value_row) else None
        out[key] = val
        mapped += 1

    core = ("project_id", "project_no", "name")
    has_core = any(out.get(k) not in (None, "") for k in core)
    non_empty_vals = sum(1 for v in out.values() if v is not None and str(v).strip() != "")

    if mapped < 2:
        if mapped == 1 and out:
            lone_k = next(iter(out.keys()))
            if lone_k in core and out[lone_k] not in (None, ""):
                return out
        return None
    if string_headers < 2:
        return None
    if has_core and non_empty_vals >= 1:
        return out
    if non_empty_vals >= 2 and mapped >= 2:
        return out
    return None


def _try_wide_project_kv_sliding(rows: List[Any], max_pairs: int = 25) -> Optional[Dict[str, Any]]:
    """横向布局表头可能不在第 1 行；表头与数据之间可能隔 1～2 个空行"""
    if len(rows) < 2:
        return None
    n_header = min(max_pairs, len(rows) - 1)
    for i in range(n_header):
        for delta in (1, 2, 3):
            j = i + delta
            if j >= len(rows):
                break
            w = _try_wide_project_kv_from_rows([rows[i], rows[j]])
            if w is not None:
                return w
    return None


def _vertical_project_key_from_excel_label(cell: str) -> Optional[str]:
    """竖向「项目」表第一列：英文内部键或导出用中文字段说明 -> 内部键"""
    s = _norm_sheet_title(cell)
    if not s:
        return None
    if s in PROJECT_KV_ENGLISH_KEYS:
        return s
    for ek in PROJECT_KV_ENGLISH_KEYS:
        if ek.casefold() == s.casefold():
            return ek
    if s in LABEL_TO_PROJECT_KEY:
        return LABEL_TO_PROJECT_KEY[s]
    for ek, lab in PROJECT_KV_FIELDS:
        if _norm_sheet_title(lab) == s:
            return ek
    return HEADER_TO_PROJECT_KEY.get(s) or HEADER_TO_PROJECT_KEY.get(s.casefold())


def _parse_kv_sheet(ws) -> Dict[str, Any]:
    """竖向「字段、值」两列；或横向「表头行+数据行」；或首列英文键竖向无表头"""
    rows_wide = list(ws.iter_rows(min_row=1, max_row=120, max_col=60, values_only=True))
    skip_wide = False
    if rows_wide and rows_wide[0] and rows_wide[0][0] is not None:
        a0 = _norm_sheet_title(str(rows_wide[0][0])).casefold()
        if a0 in ("字段", "field", "key", "名称", "属性", "项目字段"):
            skip_wide = True
    if not skip_wide:
        wide = _try_wide_project_kv_sliding(rows_wide)
        if wide is not None:
            return wide

    out: Dict[str, Any] = {}
    rows = list(ws.iter_rows(min_row=1, max_row=120, max_col=2, values_only=True))
    start_idx = 0
    if rows and rows[0] and rows[0][0] is not None:
        h = _norm_sheet_title(rows[0][0]).casefold()
        if h in ("字段", "field", "key", "名称", "属性", "项目字段"):
            start_idx = 1
    for row in rows[start_idx:]:
        if not row or row[0] is None:
            continue
        k = str(row[0]).strip()
        v = row[1] if len(row) > 1 else None
        if k:
            ik = _vertical_project_key_from_excel_label(k)
            out[ik if ik is not None else k] = v
    return out


def _cell_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s else None


def _parse_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _norm_sheet_title(s: Any) -> str:
    if s is None:
        return ""
    return (
        str(s)
        .strip()
        .replace("\u3000", "")
        .replace("\xa0", "")
    )


def _parse_cold_room_type_enum(raw: Optional[str]) -> ColdRoomType:
    s = _norm_sheet_title(str(raw or ""))
    if not s:
        return ColdRoomType.LOW_TEMP
    if s in COLD_ROOM_STR_TO_TYPE:
        return COLD_ROOM_STR_TO_TYPE[s]
    try:
        return ColdRoomType(s)
    except ValueError:
        return ColdRoomType.LOW_TEMP


def _parse_device_type_enum(raw: Optional[str]) -> DeviceType:
    s = _norm_sheet_title(str(raw or ""))
    if not s:
        return DeviceType.AIR_COOLER
    if s in DEVICE_STR_TO_TYPE:
        return DEVICE_STR_TO_TYPE[s]
    try:
        return DeviceType(s)
    except ValueError:
        return DeviceType.AIR_COOLER


def _parse_relation_type_enum(raw: Optional[str]) -> RelationType:
    s = _norm_sheet_title(str(raw or ""))
    if not s:
        return RelationType.THERMOSTAT_TO_AIR_COOLER
    if s in RELATION_STR_TO_TYPE:
        return RELATION_STR_TO_TYPE[s]
    try:
        return RelationType(s)
    except ValueError:
        return RelationType.THERMOSTAT_TO_AIR_COOLER


def _sheet_lookup(wb) -> Dict[str, str]:
    """规范化表名 -> openpyxl 原始表名（兼容大小写、空白）"""
    lu: Dict[str, str] = {}
    for raw in wb.sheetnames:
        n = _norm_sheet_title(raw)
        if not n:
            continue
        lu[n] = raw
        lu[n.casefold()] = raw
    return lu


def _open_sheet(wb, lu: Dict[str, str], *candidates: str):
    for c in candidates:
        n = _norm_sheet_title(c)
        if not n:
            continue
        if n in lu:
            return wb[lu[n]]
        if n.casefold() in lu:
            return wb[lu[n.casefold()]]
    return None


def _row_key_normalized(cell_val: Any) -> str:
    if cell_val is None:
        return ""
    s = _norm_sheet_title(cell_val).casefold().replace(" ", "").replace("_", "")
    return s.replace("（", "(").replace("）", ")")


def _fuzzy_find_project_kv_sheet(wb):
    """根据单元格出现 project_id / 项目ID 等识别项目键值表（兼容 Sheet1、键名不在 A 列）"""
    id_like = frozenset(
        {
            "projectid",
            "项目id",
            "项目编号(只读)",
            "项目编号",
        }
    )
    for raw in wb.sheetnames:
        ws = wb[raw]
        try:
            for row in ws.iter_rows(min_row=1, max_row=60, max_col=40, values_only=True):
                if not row:
                    continue
                for cell in row:
                    if cell is None:
                        continue
                    key = _row_key_normalized(cell)
                    if not key:
                        continue
                    if key == "projectid" or key in id_like:
                        return ws
        except Exception:
            continue
    return None


def _sheet_looks_like_project_kv(ws) -> bool:
    """是否像本系统导出的「项目」键值表（用于仅含 Sheet1 等单表文件）"""
    keys_hit = 0
    saw_project_no = False
    saw_name = False
    try:
        for row in ws.iter_rows(min_row=1, max_row=80, max_col=2, values_only=True):
            if not row or row[0] is None:
                continue
            raw_k = _norm_sheet_title(row[0])
            if not raw_k:
                continue
            h = raw_k.casefold()
            if h in ("字段", "field", "key", "名称", "属性", "项目字段"):
                continue
            matched = _vertical_project_key_from_excel_label(raw_k)
            if matched:
                keys_hit += 1
                if matched == "project_no":
                    saw_project_no = True
                if matched == "name":
                    saw_name = True
                continue
            kcf = raw_k.casefold()
            if kcf in PROJECT_KV_ENGLISH_KEYS:
                keys_hit += 1
                if kcf == "project_no":
                    saw_project_no = True
                if kcf == "name":
                    saw_name = True
            rk = _row_key_normalized(row[0])
            if rk == "projectid" or rk in ("项目id", "项目编号"):
                return True
        if keys_hit >= 3:
            return True
        if saw_project_no and saw_name:
            return True
        if keys_hit >= 1 and saw_name:
            return True
    except Exception:
        return False
    return False


def _fallback_single_sheet_project(wb):
    """工作簿只有一张表（常见默认名 Sheet1）时，若内容像项目键值表则采用"""
    names = [n for n in wb.sheetnames if _norm_sheet_title(n)]
    if len(names) != 1:
        return None
    ws = wb[names[0]]
    if _sheet_looks_like_project_kv(ws):
        return ws
    rows = list(ws.iter_rows(min_row=1, max_row=40, max_col=60, values_only=True))
    if _try_wide_project_kv_sliding(rows, max_pairs=35):
        return ws
    if _sheet_has_resolvable_project_header(ws):
        return ws
    return None


def _find_sheet_with_wide_project_row(wb):
    """任意工作表中是否存在「横向表头+首行数据」可解析为项目字段"""
    for raw in wb.sheetnames:
        ws = wb[raw]
        try:
            rows = list(ws.iter_rows(min_row=1, max_row=40, max_col=60, values_only=True))
            if _try_wide_project_kv_sliding(rows, max_pairs=35):
                return ws
            if _sheet_has_resolvable_project_header(ws):
                return ws
        except Exception:
            continue
    return None


def _find_project_kv_worksheet(wb):
    """定位项目键值工作表及表名索引；找不到则 (lu, None)"""
    lu = _sheet_lookup(wb)
    ws_p = _open_sheet(
        wb,
        lu,
        "项目",
        "Project",
        "project",
        "PROJECT",
        "项目信息",
        "项目主数据",
    )
    if ws_p is None:
        ws_p = _fuzzy_find_project_kv_sheet(wb)
    if ws_p is None:
        ws_p = _fallback_single_sheet_project(wb)
    if ws_p is None:
        ws_p = _find_sheet_with_wide_project_row(wb)
    if ws_p is None:
        for raw in wb.sheetnames:
            ws = wb[raw]
            try:
                if _sheet_has_resolvable_project_header(ws):
                    ws_p = ws
                    break
            except Exception:
                continue
    return lu, ws_p


def _json_safe(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return v


def extract_workbook_preview(file_bytes: bytes) -> Dict[str, Any]:
    """
    从 Excel 字节流解析结构化摘要（不写数据库）。
    工作表识别规则与「写入项目」导入一致。
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        lu, ws_p = _find_project_kv_worksheet(wb)
        if ws_p is None:
            listed = "、".join(str(x) for x in wb.sheetnames) or "(无)"
            raise ValueError(
                "无法识别项目主表：请用「导出配置」模板，或首行横向表头（如 项目名称/城市/项目编号）且下一行为数据，或竖向两列「字段|值」。"
                "下列为文件中全部工作表名："
                + listed
            )

        raw_kv = _parse_kv_sheet(ws_p)
        project_kv = {k: _json_safe(raw_kv[k]) for k in raw_kv}

        cold_rooms: List[Dict[str, Any]] = []
        ws_c = _open_sheet(wb, lu, "冷库", "ColdRoom", "cold_room", "Cold Rooms", "冷库列表")
        if ws_c is not None:
            for row in ws_c.iter_rows(min_row=2, max_row=5000, values_only=True):
                if not row or not row[0]:
                    continue
                name = str(row[0]).strip()
                if not name:
                    continue
                cold_rooms.append(
                    {
                        "冷库名称": name,
                        "room_type": _cell_str(row[1]) or "low_temp",
                        "design_temp_min": _parse_float(row[2]),
                        "design_temp_max": _parse_float(row[3]),
                        "area": _parse_float(row[4]),
                        "height": _parse_float(row[5]),
                        "volume": _parse_float(row[6]),
                        "refrigerant_type": _cell_str(row[7]),
                    }
                )

        devices: List[Dict[str, Any]] = []
        ws_d = _open_sheet(wb, lu, "设备", "Device", "Devices", "设备清单", "devices")
        if ws_d is not None:
            for row in ws_d.iter_rows(min_row=2, max_row=10000, values_only=True):
                if not row or not row[0]:
                    continue
                device_no = str(row[0]).strip()
                if not device_no:
                    continue
                devices.append(
                    {
                        "device_no": device_no,
                        "device_type": _cell_str(row[1]) or "air_cooler",
                        "cold_room_name": _cell_str(row[2]),
                        "brand": str(row[3] or "").strip() or None,
                        "model": str(row[4] or "").strip() or None,
                        "defrost_method": _cell_str(row[5]),
                        "has_intelligent_defrost": _cell_str(row[6]),
                        "expansion_valve_type": _cell_str(row[7]),
                        "factory_no": _cell_str(row[8]),
                        "comm_port_type": _cell_str(row[9]),
                        "comm_protocol": _cell_str(row[10]),
                        "gateway_id": _parse_int(row[11]),
                        "gateway_port": _parse_int(row[12]),
                        "rs485_address": _cell_str(row[13]),
                        "specifications": _cell_str(row[14]),
                        "remarks": _cell_str(row[15]),
                    }
                )

        relations: List[Dict[str, Any]] = []
        ws_r = _open_sheet(
            wb,
            lu,
            "设备关系",
            "DeviceRelation",
            "device_relations",
            "Relations",
            "relations",
            "关系",
        )
        if ws_r is not None:
            for row in ws_r.iter_rows(min_row=2, max_row=10000, values_only=True):
                if not row or not row[0] or not row[1]:
                    continue
                relations.append(
                    {
                        "from_device_no": str(row[0]).strip(),
                        "to_device_no": str(row[1]).strip(),
                        "relation_type": _cell_str(row[2]) or "thermostat_to_air_cooler",
                        "description": _cell_str(row[3]),
                    }
                )

        return {
            "sheet_names": list(wb.sheetnames),
            "project": project_kv,
            "cold_rooms": cold_rooms,
            "devices": devices,
            "relations": relations,
            "summary": {
                "cold_room_count": len(cold_rooms),
                "device_count": len(devices),
                "relation_count": len(relations),
            },
        }
    finally:
        wb.close()


def apply_import(db: Session, project_id: int, file_bytes: bytes) -> Dict[str, int]:
    """将 Excel 配置合并到指定项目（冷库/设备按名称或编号 upsert；关系整表替换）。"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise ValueError("项目不存在")

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lu, ws_p = _find_project_kv_worksheet(wb)
    if ws_p is None:
        listed = "、".join(str(x) for x in wb.sheetnames) or "(无)"
        raise ValueError(
            "缺少项目主表：请使用「导出配置」模板，或表名「项目」/「Project」，"
            "或横向表头（项目名称、城市、项目编号等）与数据行（中间可隔 1～2 空行），或竖向 project_id/name 键值。"
            "下列为文件中全部工作表名：" + listed
        )
    kv = _parse_kv_sheet(ws_p)
    pid_in_file = kv.get("project_id")
    if pid_in_file is not None and pid_in_file != "":
        try:
            pid_parsed = int(float(pid_in_file))
        except (TypeError, ValueError):
            pid_parsed = None
        if pid_parsed is not None and pid_parsed != int(project_id):
            raise ValueError(f"文件中 project_id={pid_in_file} 与当前项目 {project_id} 不一致，禁止导入")

    # 更新项目（不修改 project_no / id）
    if kv.get("name") is not None:
        project.name = str(kv.get("name") or "").strip() or project.name
    for attr in (
        "end_customer",
        "business_type",
        "city",
        "address",
        "mailing_address",
        "recipient_name",
        "recipient_phone",
        "remarks",
    ):
        if attr in kv and kv[attr] is not None:
            val = kv[attr]
            setattr(project, attr, str(val).strip() if val is not None else None)

    if "expected_arrival_time" in kv and kv["expected_arrival_time"] not in (None, ""):
        raw = kv["expected_arrival_time"]
        if isinstance(raw, datetime):
            project.expected_arrival_time = raw.date()
        else:
            s = str(raw).strip()[:10]
            project.expected_arrival_time = date.fromisoformat(s)

    n_cold = n_dev = n_rel = 0

    # 冷库
    ws_c = _open_sheet(wb, lu, "冷库", "ColdRoom", "cold_room", "Cold Rooms", "冷库列表")
    if ws_c is not None:
        rows = list(ws_c.iter_rows(min_row=2, values_only=True))
        name_to_room: Dict[str, ColdRoom] = {
            r.name: r
            for r in db.query(ColdRoom).filter(ColdRoom.project_id == project_id).all()
        }
        for row in rows:
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            if not name:
                continue
            rt_raw = _cell_str(row[1]) or "low_temp"
            room_type = _parse_cold_room_type_enum(rt_raw)
            area = _parse_float(row[4])
            height = _parse_float(row[5])
            vol = _parse_float(row[6])
            if vol is None and area and height:
                vol = round(area * height, 2)
            payload = {
                "room_type": room_type,
                "design_temp_min": _parse_float(row[2]),
                "design_temp_max": _parse_float(row[3]),
                "area": area,
                "height": height,
                "volume": vol,
                "refrigerant_type": _cell_str(row[7]),
            }
            if name in name_to_room:
                cr = name_to_room[name]
                for k, v in payload.items():
                    setattr(cr, k, v)
            else:
                cr = ColdRoom(project_id=project_id, name=name, **payload)
                db.add(cr)
                name_to_room[name] = cr
            n_cold += 1
        db.flush()
        # 刷新 id 映射
        name_to_id = {
            r.name: r.id
            for r in db.query(ColdRoom).filter(ColdRoom.project_id == project_id).all()
        }
    else:
        name_to_id = {
            r.name: r.id
            for r in db.query(ColdRoom).filter(ColdRoom.project_id == project_id).all()
        }

    # 设备
    ws_d = _open_sheet(wb, lu, "设备", "Device", "Devices", "设备清单", "devices")
    if ws_d is not None:
        rows = list(ws_d.iter_rows(min_row=2, values_only=True))
        no_to_dev: Dict[str, Device] = {
            d.device_no: d
            for d in db.query(Device).filter(Device.project_id == project_id).all()
        }
        for row in rows:
            if not row or not row[0]:
                continue
            device_no = str(row[0]).strip()
            if not device_no:
                continue
            dt_raw = _cell_str(row[1]) or "air_cooler"
            dtype = _parse_device_type_enum(dt_raw)
            cr_name = _cell_str(row[2])
            cold_id = name_to_id.get(cr_name) if cr_name else None

            fields = {
                "device_type": dtype,
                "cold_room_id": cold_id,
                "brand": str(row[3] or "").strip() or "-",
                "model": str(row[4] or "").strip() or "-",
                "defrost_method": _cell_str(row[5]),
                "has_intelligent_defrost": _cell_str(row[6]),
                "expansion_valve_type": _cell_str(row[7]),
                "factory_no": _cell_str(row[8]),
                "comm_port_type": _cell_str(row[9]),
                "comm_protocol": _cell_str(row[10]),
                "gateway_id": _parse_int(row[11]),
                "gateway_port": _parse_int(row[12]),
                "rs485_address": _cell_str(row[13]),
                "specifications": _cell_str(row[14]),
                "remarks": _cell_str(row[15]),
            }
            if device_no in no_to_dev:
                d = no_to_dev[device_no]
                for k, v in fields.items():
                    setattr(d, k, v)
            else:
                d = Device(project_id=project_id, device_no=device_no, **fields)
                db.add(d)
                no_to_dev[device_no] = d
            n_dev += 1
        db.flush()
    else:
        no_to_dev = {
            d.device_no: d
            for d in db.query(Device).filter(Device.project_id == project_id).all()
        }

    # 关系：整表替换
    ws_r = _open_sheet(
        wb,
        lu,
        "设备关系",
        "DeviceRelation",
        "device_relations",
        "Relations",
        "relations",
        "关系",
    )
    if ws_r is not None:
        db.query(DeviceRelation).filter(DeviceRelation.project_id == project_id).delete(synchronize_session=False)
        db.flush()
        rows = list(ws_r.iter_rows(min_row=2, values_only=True))
        # 重新加载设备编号 -> id
        no_to_id = {
            d.device_no: d.id
            for d in db.query(Device).filter(Device.project_id == project_id).all()
        }
        for row in rows:
            if not row or not row[0] or not row[1]:
                continue
            fno = str(row[0]).strip()
            tno = str(row[1]).strip()
            rt_raw = _cell_str(row[2]) or "thermostat_to_air_cooler"
            rtype = _parse_relation_type_enum(rt_raw)
            fid = no_to_id.get(fno)
            tid = no_to_id.get(tno)
            if not fid or not tid:
                continue
            db.add(
                DeviceRelation(
                    project_id=project_id,
                    from_device_id=fid,
                    to_device_id=tid,
                    relation_type=rtype,
                    description=_cell_str(row[3]),
                )
            )
            n_rel += 1

    wb.close()
    return {"cold_rooms": n_cold, "devices": n_dev, "relations": n_rel}
